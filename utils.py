import zipfile
import os
import pandas as pd
from bs4 import BeautifulSoup


# 📦 Unzip file
def unzip_file(zip_path, extract_to):
    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
        zip_ref.extractall(extract_to)
    print(f"✅ Extracted: {zip_path}")


def extract_full_summary(html_file):
    import pandas as pd
    from bs4 import BeautifulSoup

    with open(html_file, "r", encoding="utf-8", errors="ignore") as f:
        soup = BeautifulSoup(f, "lxml")

  
    table = soup.find("table", {"id": "TransactionsTable"})

    transactions_df = None

    if table:
        rows = []
        for tr in table.find_all("tr"):
            cols = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if cols:
                rows.append(cols)

        if len(rows) > 1:
            transactions_df = pd.DataFrame(rows[1:], columns=rows[0])
            transactions_df.columns = transactions_df.columns.str.strip()

    
    summary_data = {}

    # 🔥 Extract PERIOD
    full_text = soup.get_text("\n")
    for line in full_text.split("\n"):
        if line.strip().startswith("Period:"):
            summary_data["Period"] = line.replace("Period:", "").strip()

    # 🔥 Extract ALL <td> values
    tds = [td.get_text(strip=True) for td in soup.find_all("td")]

    # 🎯 Keys we want
    stats_keys = [
        "Maximum Running Vusers",
        "Total Throughput (bytes)",
        "Average Throughput (B/s)",
        "Total Hits:",
        "Average Hits per Second",
        "Passed Transactions Ratio"
    ]

    # 🔥 Pair KEY → VALUE (td[i] → td[i+1])
    for i in range(len(tds) - 1):
        key = tds[i]
        value = tds[i + 1]

        if key in stats_keys:
            summary_data[key.replace(":", "")] = value

    # 📊 Convert to DataFrame
    summary_df = pd.DataFrame(
        list(summary_data.items()),
        columns=["Metric", "Value"]
    )

    return transactions_df, summary_df


# 📊 Create final Excel with multiple sheets
def create_full_excel(report_folder, output_file):
    html_file = os.path.join(report_folder, "summary.html")

    if not os.path.exists(html_file):
        print("❌ summary.html not found")
        return

    transactions_df, summary_df = extract_full_summary(html_file)

    if transactions_df is None:
        print("❌ No transaction table found")
        return

    # 📂 Ensure output folder exists
    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # 📊 Write to Excel
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        transactions_df.to_excel(writer, sheet_name="Transactions", index=False)

        if not summary_df.empty:
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

    print(f"🎯 Full Excel created: {output_file}")

#Compare multiple Excels and add trend analysis
def compare_multiple_excels(files, output_file):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font

    merged_df = None

    for i, file in enumerate(files):
        df = pd.read_excel(file)

        if "Transaction Name" not in df.columns:
            continue

        cols_to_keep = ["Transaction Name"]

        for col in ["Average", "Minimum", "Maximum", "Pass", "Fail"]:
            if col in df.columns:
                cols_to_keep.append(col)

        df = df[cols_to_keep]

        # Rename columns
        rename_map = {
            col: f"{col}_file{i+1}"
            for col in df.columns if col != "Transaction Name"
        }
        df = df.rename(columns=rename_map)

        if merged_df is None:
            merged_df = df
        else:
            merged_df = pd.merge(
                merged_df,
                df,
                on="Transaction Name",
                how="outer"
            )

    if merged_df is None:
        print("❌ No valid data to compare")
        return

    # =============================
    # 🔥 ADD TREND COLUMN
    # =============================
    avg_cols = [col for col in merged_df.columns if "Average_file" in col]

    if len(avg_cols) >= 2:
        first_col = avg_cols[0]
        last_col = avg_cols[-1]

        def get_trend(row):
            try:
                first = float(row[first_col])
                last = float(row[last_col])

                if last < first:
                    return "⬆ Improved"
                elif last > first:
                    return "⬇ Degraded"
                else:
                    return "➖ No Change"
            except:
                return "N/A"

        merged_df["Trend"] = merged_df.apply(get_trend, axis=1)

    # =============================
    # 💾 SAVE
    # =============================
    merged_df.to_excel(output_file, index=False)

    # =============================
    # 🎨 APPLY COLOR FORMATTING
    # =============================
    workbook = load_workbook(output_file)
    sheet = workbook.active
    
    # Define colors
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    red_font = Font(color="FFFFFF", bold=True)
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    green_font = Font(color="FFFFFF", bold=True)
    
    # Find Trend column
    trend_col = None
    for col_idx, cell in enumerate(sheet[1], 1):
        if cell.value == "Trend":
            trend_col = col_idx
            break
    
    # Apply colors to Trend column cells
    if trend_col:
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=trend_col)
            if cell.value:
                if "Degraded" in str(cell.value):
                    cell.fill = red_fill
                    cell.font = red_font
                elif "Improved" in str(cell.value):
                    cell.fill = green_fill
                    cell.font = green_font
    
    workbook.save(output_file)

    print(f"📊 Comparison with trend created: {output_file}")
#send email with attachment using Outlook COM
def send_email_outlook(subject, body, to, attachment_path, sender=None):
    import win32com.client

    outlook = win32com.client.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)

    mail.To = to
    mail.Subject = subject
    mail.Body = body

    # 🔥 SET SENDER ACCOUNT
    if sender:
        for account in outlook.Session.Accounts:
            if account.SmtpAddress.lower() == sender.lower():
                mail._oleobj_.Invoke(*(64209, 0, 8, 0, account))  # magic line
                break

    if isinstance(attachment_path, (list, tuple)):
        for path in attachment_path:
            if path and os.path.exists(path):
                mail.Attachments.Add(path)
    elif attachment_path and os.path.exists(attachment_path):
        mail.Attachments.Add(attachment_path)

    mail.Send()

    print(f"📧 Email sent from {sender} to {to}")